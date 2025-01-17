from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from io import BytesIO
from PIL import Image as PILImage

def excel_to_pdf(excel_path, pdf_path):
    """
    Excel(.xlsx) 内の全シートを自動で読み取り、
    セルの表データと画像を横向き(A4)の PDF に出力するサンプル関数
    """
    
    # 1) Excelファイルを読み込む (data_only=True で数式セルは値のみ取得)
    workbook = load_workbook(excel_path, data_only=True)
    
    # 2) ReportLab で PDF を作るための準備 (A4 を横向きに設定)
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A4),   # A4を横向きに
    )
    
    # PDF に配置する「要素(ストーリー)」を積み上げていくリスト
    story = []
    
    # デザイン用スタイル(見出しなど)を取得
    styles = getSampleStyleSheet()
    
    # 3) すべてのシートをループ
    for sheetname in workbook.sheetnames:
        sheet = workbook[sheetname]
        
        # --- (A) シート名の見出し ---
        heading = Paragraph(f"Sheet: {sheetname}", styles["Heading2"])
        story.append(heading)
        story.append(Spacer(1, 0.5 * cm))  # 少し空行を入れる
        
        # --- (B) セルの表データを取得して Table で配置 ---
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # 例: (A1 から最終行・最終列まで) を values_only=True で取得
        data = []
        for row in sheet.iter_rows(
            min_row=1, max_row=max_row,
            min_col=1, max_col=max_col,
            values_only=True
        ):
            # row はタプルになるので、リストに変換して積む
            data.append(list(row))
        
        if data:  # シートによっては空の可能性もあるため
            table = Table(data)
            # 表に枠線などのスタイルを指定
            table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.5 * cm))
        
        # --- (C) シート内の画像を取得して配置 ---
        # openpyxl の画像オブジェクトは sheet._images という非公開属性に格納される
        for img_obj in getattr(sheet, '_images', []):
            # 画像データ(バイナリ)を取得
            # openpyxl.drawing.image.Image オブジェクト: 例) img_obj._data
            # バージョンによっては img_obj._data() または img_obj._data プロパティなど異なる場合があるので注意
            img_data = img_obj._data()
            
            # Pillow で一度読み込む
            pil_img = PILImage.open(BytesIO(img_data))
            
            # PDF で扱えるように BytesIO に書き込んで ReportLab の Image に渡す
            tmp_io = BytesIO()
            pil_img.save(tmp_io, format='PNG')  # PNG などに変換
            tmp_io.seek(0)
            
            # Platypus の Image オブジェクトとして追加
            rl_img = Image(tmp_io)
            
            # 大きすぎる場合などは必要に応じてサイズを縮小する
            # 例: 幅を 15cm にする (縦横比は維持)
            desired_width = 15 * cm
            ratio = desired_width / pil_img.width
            rl_img._restrictSize(desired_width, pil_img.height * ratio)
            
            story.append(rl_img)
            story.append(Spacer(1, 0.5 * cm))

        # シートごとに改ページの代わり
        story.append(Spacer(1, 1 * cm))

    # 4) ドキュメントビルド (PDF ファイル生成)
    doc.build(story)


if __name__ == "__main__":
    excel_file = ".xls"   # 読み込みたい Excel ファイルパス
    pdf_file = ".pdf"      # 保存したい PDF ファイルパス
    
    excel_to_pdf(excel_file, pdf_file)
    print(f"PDF 生成完了: {pdf_file}")

