# -*- coding: utf-8 -*-
import sys
from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import cm
from io import BytesIO
from PIL import Image as PILImage

# ▼ 日本語フォントを使う場合（文字化けや豆腐対策）
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
pdfmetrics.registerFont(TTFont('MS-Gothic', 'msgothic.ttc'))

def excel_to_pdf_keep_images_near(excel_path, pdf_path):
    """
    Excel(.xlsx) 内の全シートを処理し、
    ・各シートの行データを1行ずつテーブル化
    ・画像はアンカー位置の行の直後に配置
    することで、画像とデータが離れにくいPDFを生成するサンプル
    """
    
    # Excel読み込み
    wb = load_workbook(excel_path, data_only=True)
    
    # PDFドキュメント設定
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A3)
    )
    story = []
    styles = getSampleStyleSheet()
    
    # 日本語フォントを表や見出しに適用
    styles["Normal"].fontName = 'MS-Gothic'
    styles["Heading2"].fontName = 'MS-Gothic'
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # --- シート見出し ---
        heading_txt = f"シート: {sheet_name}"
        story.append(Paragraph(heading_txt, styles["Heading2"]))
        story.append(Spacer(1, 0.5 * cm))
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # ========== 画像のアンカーを取得して、行番号に紐づける ========== #
        images_in_row = {}
        
        # sheet._images は非公開属性なのでバージョンによって異なる可能性があることに注意
        for img_obj in getattr(sheet, "_images", []):
            anchor = img_obj.anchor  # 例: "C5" など
            # anchor から行番号(row)と列番号(col)を取得
            if isinstance(anchor, str):
                # "C5" のような文字列の場合
                coord = coordinate_from_string(anchor)  # ('C', '5') など
                colx = column_index_from_string(coord[0])  # C -> 3
                rowx = int(coord[1])                     # 5
            else:
                # anchor が OneCellAnchor / TwoCellAnchor オブジェクトの場合
                # openpyxlのバージョンによって取り方が違う
                # 例: anchor._from.row_idx など
                # ここでは仮に rowx = anchor._from.row_idx + 1 のような形で取得
                rowx = anchor._from.row_idx + 1
                colx = anchor._from.col_idx + 1
            
            # images_in_row[row] に格納
            if rowx not in images_in_row:
                images_in_row[rowx] = []
            images_in_row[rowx].append(img_obj)
        
        # ========== 行データを順次PDFに配置 ========== #
        for r in range(1, max_row + 1):
            # その行のセル値をリスト化
            row_values = []
            for c in range(1, max_col + 1):
                val = sheet.cell(row=r, column=c).value
                row_values.append("" if val is None else val)
            
            # その行に実際に値がなかったらスキップしたい場合は以下をコメントイン
            # if all(cell == "" for cell in row_values):
            #     continue
            
            # この1行だけのTableを作成
            table_data = [row_values]  # 1行
            row_table = Table(table_data)
            row_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTNAME', (0,0), (-1,-1), 'MS-Gothic'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            story.append(row_table)
            
            # もしこの行に画像があれば、行の直後に画像を配置
            if r in images_in_row:
                for img_obj in images_in_row[r]:
                    # 画像バイナリ
                    img_data = img_obj._data()
                    pil_img = PILImage.open(BytesIO(img_data))
                    
                    tmp_io = BytesIO()
                    pil_img.save(tmp_io, format='PNG')
                    tmp_io.seek(0)
                    
                    rl_img = Image(tmp_io)
                    
                    # 幅を15cmなどに制限（お好みで）
                    desired_width = 15 * cm
                    ratio = desired_width / pil_img.width
                    rl_img._restrictSize(desired_width, pil_img.height * ratio)
                    
                    story.append(rl_img)
                    
            # 行ごとに少し余白
            story.append(Spacer(1, 0.3 * cm))
        
        # シートごとに大きめに区切りを入れる
        story.append(Spacer(1, 1 * cm))
    
    # PDFに書き出し
    doc.build(story)
    print(f"PDF 生成完了: {pdf_path}")


if __name__ == "__main__":
    excel_file = "サンプル.xlsx"
    pdf_file   = "出力サンプル_A3_画像近接配置.pdf"
    excel_to_pdf_keep_images_near(excel_file, pdf_file)

