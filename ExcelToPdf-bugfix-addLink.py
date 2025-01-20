# -*- coding: utf-8 -*-
import sys
import os
import glob
from io import BytesIO
from urllib.parse import quote  # ExcelファイルパスをURLエンコードするため
from PIL import Image as PILImage

# ▼ Windowsコンソールの文字化け対策（PowerShellの場合は "chcp 65001" が推奨）
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle, Image, KeepInFrame
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import cm

# ▼ 日本語フォントを使う（Windows標準のMSゴシックを例に）
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
pdfmetrics.registerFont(TTFont('MS-Gothic', 'msgothic.ttc'))


def excel_to_pdf_keep_images_near(excel_path, pdf_path):
    """
    指定した Excel(.xlsx) ファイルを読み込み、以下を行う:
      - 最初のページの先頭に「元ファイルへのリンク」を追加
      - 全シートを処理（シート名を見出しとして出力）
      - 各シートを行単位で出力し、同じ行に画像があれば直後に配置
      - A3横向きでPDFを作成
      - "too large on page X..." エラーを回避するために KeepInFrame + 画像縮小
      - Windows上でも日本語文字が文字化け/豆腐にならないようにMSゴシックを使用
    """
    # (1) Excel読み込み
    wb = load_workbook(excel_path, data_only=True)
    
    # (2) PDFドキュメント(A3横向き)の準備
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A3)
    )
    story = []
    styles = getSampleStyleSheet()
    
    # (3) 日本語フォント設定
    styles["Normal"].fontName = 'MS-Gothic'
    styles["Heading2"].fontName = 'MS-Gothic'
    
    # --- (A) 最初のページに「元ファイルへのリンク」を追加 ---
    # file:/// (ローカルファイルパス) をURLエンコードしてハイパーリンクにする
    abs_excel_path = os.path.abspath(excel_path)
    excel_file_url = "file:///" + quote(abs_excel_path.replace("\\", "/"))
    
    link_text = f'<link href="{excel_file_url}">元ファイル: {os.path.basename(excel_path)}</link>'
    link_paragraph = Paragraph(link_text, styles["Normal"])
    story.append(link_paragraph)
    story.append(Spacer(1, 1 * cm))  # 余白
    
    # ========== (B) シートごとの出力 ==========
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # --- シート見出し ---
        heading_txt = f"シート: {sheet_name}"
        heading = Paragraph(heading_txt, styles["Heading2"])
        story.append(heading)
        story.append(Spacer(1, 0.5 * cm))
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # ========== (1) 画像アンカーを行番号に紐づける ==========
        images_in_row = {}
        for img_obj in getattr(sheet, "_images", []):
            anchor = img_obj.anchor
            rowx = None
            
            # パターン1) anchor が "C5" のような文字列
            if isinstance(anchor, str):
                coord = coordinate_from_string(anchor)  # 例: ('C','5')
                rowx = int(coord[1])
                # colx = column_index_from_string(coord[0])
            else:
                # パターン2) OneCellAnchor / TwoCellAnchor オブジェクトの場合
                # バージョンによってプロパティが変わるため要注意
                # 例: anchor._from.row_idx + 1
                rowx = anchor._from.row_idx + 1
            
            if rowx:
                if rowx not in images_in_row:
                    images_in_row[rowx] = []
                images_in_row[rowx].append(img_obj)
        
        # ========== (2) 各行ごとにテーブル → 画像 の順で追加 ==========
        for r in range(1, max_row + 1):
            # 行データ
            row_values = []
            for c in range(1, max_col + 1):
                val = sheet.cell(row=r, column=c).value
                row_values.append("" if val is None else str(val))
            
            # 1行N列の Table を作成
            row_table = Table([row_values])
            row_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTNAME', (0,0), (-1,-1), 'MS-Gothic'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            
            # Table が巨大になる場合を想定し、KeepInFrameで縮小対応
            wrapped_table = KeepInFrame(
                width=38*cm, height=20*cm,
                content=[row_table],
                mode='shrink'
            )
            story.append(wrapped_table)
            
            # この行に画像があれば、続けて配置
            if r in images_in_row:
                for img_obj in images_in_row[r]:
                    # 画像バイナリ
                    if callable(img_obj._data):
                        img_data = img_obj._data()
                    else:
                        img_data = img_obj._data
                    pil_img = PILImage.open(BytesIO(img_data))
                    
                    # ReportLab Image 化
                    tmp_io = BytesIO()
                    pil_img.save(tmp_io, format='PNG')
                    tmp_io.seek(0)
                    
                    rl_img = Image(tmp_io)
                    
                    # 幅15cmに縮小しつつ縦横比を維持
                    desired_width = 15 * cm
                    ratio = desired_width / pil_img.width
                    rl_img._restrictSize(desired_width, pil_img.height * ratio)
                    
                    # 画像もフレームオーバー時に縮小対応
                    wrapped_img = KeepInFrame(
                        width=38*cm, height=20*cm,
                        content=[rl_img],
                        mode='shrink'
                    )
                    story.append(wrapped_img)
            
            # 各行間に少し余白
            story.append(Spacer(1, 0.3 * cm))
        
        # シート区切り
        story.append(Spacer(1, 1 * cm))
    
    # (C) PDFビルド
    doc.build(story)
    print(f"[完了] {excel_path} -> {pdf_path}")


def convert_folder_excel_to_pdf(folder_path):
    """
    指定フォルダ内の .xlsx ファイルをすべて検索し、
    excel_to_pdf_keep_images_near() でPDFへ変換
    """
    folder_path = os.path.abspath(folder_path)
    xlsx_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    
    if not xlsx_files:
        print("指定フォルダに .xlsx ファイルが見つかりませんでした。")
        return
    
    print(f"変換対象ファイル数: {len(xlsx_files)}")
    
    for excel_file in xlsx_files:
        # 同名の .pdf に変換
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        pdf_name = base_name + ".pdf"
        pdf_path = os.path.join(folder_path, pdf_name)
        
        excel_to_pdf_keep_images_near(excel_file, pdf_path)

    print("フォルダ内の全ファイル変換が完了しました。")


if __name__ == "__main__":
    # 例: C:/path/to/excel_folder
    target_folder = r"C:\Users\YourName\Documents\excel_folder"
    convert_folder_excel_to_pdf(target_folder)



