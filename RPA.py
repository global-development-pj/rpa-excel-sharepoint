# -*- coding: utf-8 -*-
"""
Created on Sat Jan 11 11:26:58 2025

@author: DELL
"""

import gspread
from oauth2client.service_account import ServiceAccountCredentials
import openpyxl
import pandas as pd
from msal import ConfidentialClientApplication
import os
import requests
import config
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential

class ShoppingRPA:
    def convert_data_to_required_format(self, dataframe, append_data=False):
        val_list = None
        if append_data==False:
            val_list = [dataframe.columns.tolist()]
            val_list += dataframe.values.tolist()
        else:
            val_list = dataframe.values.tolist()
        return val_list
    
    def upload_to_sharepoint(self):
        client_id = config['ClientId']
        tenant_id = config['TenantId']
        tenant = config['tenant']
        authority = f"https://login.microsoftonline.com/{tenant_id}"
        scope = [f"https://{tenant}.sharepoint.com/.default"]
        username = config['sp_user']
        password = config['sp_password']
        sitename = "rpa"
        site_url = "https://globaldevelopmentpj.sharepoint.com/sites/rpa"
        file_name = "file+date..xlsx"
        local_file_path = os.path.abspath(file_name)
        target_folder_url = "/sites/rpa/Shared Documents/ExcelPython"
        # Authenticate with SharePoint
        credentials = UserCredential(username, password)
        ctx = ClientContext(site_url).with_credentials(credentials)
        try:
            # Get the target folder
            target_folder = ctx.web.get_folder_by_server_relative_url(target_folder_url)
            # Upload the file
            with open(local_file_path, 'rb') as file:
                # Get the file name
                file_name = os.path.basename(local_file_path)
                # Upload the file
                target_folder.upload_file(file_name, file.read()).execute_query()
                print(f"File '{file_name}' uploaded successfully to '{target_folder_url}'")
        except Exception as e:
            print(f"An error occurred: {e}")
    
    def write_to_spreadsheet(self, file_name, data, sheet_name, append_data=False):
        
        Sheet_Name_Dict = {"A":0, "B":1, "C":2, "D": 3}
        sheet_no = Sheet_Name_Dict[sheet_name]
        scopes = [
            'https://www.googleapis.com/auth/spreedsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        
        spreedsheet_name = file_name.split('.')[0]
        creds = ServiceAccountCredentials.from_json_keyfile_name('secret_key.json', scopes=scopes)
        file = gspread.authorize(creds)
        
        workbook = file.open(spreedsheet_name)
        sheet = workbook.get_worksheet(sheet_no)
        
        if append_data == False:
            row_no = 1
            max_row_val = sheet.cell(row=row_no, col=1).value
            while max_row_val!=None:
                for col_no in range(0, 4):
                    sheet.cell(row=row_no, col=col_no).value = None
                row_no+=1
                max_row_val = sheet.cell(row=row_no, col=1).value
            
            row_no=1
            for row in data:
                for col_no in range(0, len(row)):
                    sheet.cell(row=row_no, col=col_no).value = None
                row_no+=1
                
        elif append_data == True:
            row_no = 1
            max_row_val = sheet.cell(row=row_no, col=1).value
            while max_row_val!=None:
                row_no+=1
                max_row_val = sheet.cell(row=row_no, col=1).value
            
            for row in data:
                for col_no in range(0, len(row)):
                    sheet.cell(row=row_no, col=col_no).value = row[col_no]
                row_no+=1
    
    def save_tocsv_local(self, infile, outfile, data, sheet_name, append_data=False):
        workbook = openpyxl.load_workbook(infile)
        sheet = workbook[sheet_name]
        if append_data==True:
            row_no = sheet.max_row
            for row in data:
                row_no += 1
                for colno in range(1, len(row)+1):
                    sheet.cell(row_no, colno).value = row[colno-1]
        else:
            row_no = 0
            for row in data:
                row_no += 1
                for colno in range(1, len(row)+1):
                    sheet.cell(row_no, colno).value = row[colno-1]
        workbook.save(outfile)
        
    def delete_connection(self, filename):
        workbook = openpyxl.load_workbook(filename)
        sheet = workbook['F']
        for colno in range(1, 5):
            sheet.cell(1, colno).value = ''
        workbook.save(filename)
    
    def extract_data_from_RDS(self, conn_str, sql_query):
        data = pd.read_sql(sql_query, conn_str)
        return data
    
    def convert_date_to_12hour_clock(self, dataframe, col_name):
        dates = []
        for row in dataframe[col_name]:
            #dates.append(row.strftime('%m/%d/%y %H:%M %p'))
            dates.append(row.strftime('%Y%m%d'))
        temp_df = pd.DataFrame({col_name:dates})
        dataframe[col_name] = temp_df[col_name]
        return dataframe

    def main(self, infile, outfile, conn_str, sheet_name, query, date_cols=[], append_data=False):
        dataframe = self.extract_data_from_RDS(conn_str, query)
        for col_name in date_cols:
            dataframe = self.convert_date_to_12hour_clock(dataframe, col_name)
        values = self.convert_data_to_required_format(dataframe, append_data=append_data)
        self.save_tocsv_local(infile, outfile, values, sheet_name, append_data)
        self.write_to_spreadsheet(outfile, values, sheet_name, append_data)
        self.delete_connection_local(outfile)
    
run_parameters = [{'infile':'Book_master.xlsx','outfile':'file+date..xlsx',
                   'conn_str':'postgresql://powerautomate:powerautomate-135@terraform-20241206145403139300000001.cd228oem44w3.us-east-1.rds.amazonaws.com:5432/powerdb',
                  'query':"SELECT u.username, u.email, o.product_name, o.order_date FROM users u INNER JOIN orders o ON u.user_id = o.user_id WHERE o.order_date <= CAST(NOW() AS DATE);",
                  'date_cols':['order_date'], 'append_data':True, 'Sheet_name':'A'},
                  {'infile':'file+date1..xlsx','outfile':'file+date..xlsx',
                   'conn_str':'postgresql://powerautomate:powerautomate-135@terraform-20241206145403139300000001.cd228oem44w3.us-east-1.rds.amazonaws.com:5432/powerdb',
                    'query':"SELECT * FROM orders;", 'date_cols':['order_date'], 'append_data':False, 'Sheet_name':'B'},
                  {'infile':'file+date1..xlsx','outfile':'file+date..xlsx',
                   'conn_str':'postgresql://powerautomate:powerautomate-135@terraform-20241206145403139300000001.cd228oem44w3.us-east-1.rds.amazonaws.com:5432/powerdb',
                   'query':"SELECT * FROM USERS;", 'date_cols':['created_at'], 'append_data':False, 'Sheet_name':'C'},
                  {'infile':'file+date1..xlsx','outfile':'file+date..xlsx',
                   'conn_str':'postgresql://powerautomate:powerautomate-135@terraform-20241206145403139300000001.cd228oem44w3.us-east-1.rds.amazonaws.com:5432/powerdb',
                    'query':"SELECT u.username, u.email, o.product_name, o.order_date FROM users u INNER JOIN orders o ON u.user_id = o.user_id", 
                    'date_cols':['order_date'], 'append_data':False, 'Sheet_name':'D'}
                ]
shopping = ShoppingRPA()

for parameters in run_parameters:
    shopping.main(parameters['infile'], parameters['outfile'], parameters['conn_str'], parameters['Sheet_name'], parameters['query'], parameters['date_cols'], parameters['append_data'])



#def save_tocsv_local(infile, outfile, values, sheet_name, append_data):   