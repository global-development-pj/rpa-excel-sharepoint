# -*- coding: utf-8 -*-
import sys
# 必要に応じてコンソールの文字化け対策
# sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import cm
from io import BytesIO
from PIL import Image as PILImage

# ▼ 日本語フォントを使う場合は追加登録が必要（後述）
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 例: Windows の「MSゴシック」を使う場合のフォント登録
pdfmetrics.registerFont(TTFont('MS-Gothic', 'msgothic.ttc'))

def excel_to_pdf(excel_path, pdf_path):
    """
    指定した Excel(.xlsx) ファイル内の全シートを読み取り、
    A3・横向きの PDF にデータ(セル値)と画像を出力する
    (Windows上でも文字化けしにくいようにフォントをMSゴシックに)
    """
    
    workbook = load_workbook(excel_path, data_only=True)
    
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A3)
    )
    
    story = []
    styles = getSampleStyleSheet()
    
    # 例: デフォルトスタイルを上書きして日本語フォントに変更
    #     (これをしないとPDF内が□(豆腐)表示になる場合がある)
    styles["Normal"].fontName = 'MS-Gothic'
    styles["Heading2"].fontName = 'MS-Gothic'
    
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        heading_text = f"シート名: {sheet_name}"
        heading = Paragraph(heading_text, styles["Heading2"])
        story.append(heading)
        story.append(Spacer(1, 0.5 * cm))
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row,
                                   min_col=1, max_col=max_col,
                                   values_only=True):
            row_list = [(cell if cell is not None else "") for cell in row]
            data.append(row_list)
        
        if data:
            table = Table(data)
            table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTNAME', (0,0), (-1,-1), 'MS-Gothic'),  # 日本語フォント
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            story.append(table)
            story.append(Spacer(1, 0.5 * cm))
        
        # 画像を取得して配置
        for img_obj in getattr(sheet, '_images', []):
            img_data = img_obj._data()  # バージョンによっては _data など
            pil_img = PILImage.open(BytesIO(img_data))
            
            tmp_io = BytesIO()
            pil_img.save(tmp_io, format='PNG')
            tmp_io.seek(0)
            
            rl_img = Image(tmp_io)
            
            desired_width = 20 * cm
            ratio = desired_width / pil_img.width
            rl_img._restrictSize(desired_width, pil_img.height * ratio)
            
            story.append(rl_img)
            story.append(Spacer(1, 0.5 * cm))
        
        story.append(Spacer(1, 1 * cm))
    
    doc.build(story)
    print(f"PDF 生成完了: {pdf_path}")


if __name__ == "__main__":
    excel_file = "サンプル.xlsx"  
    pdf_file = "出力サンプル_A3.pdf"
    
    excel_to_pdf(excel_file, pdf_file)

