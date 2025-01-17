from openpyxl import load_workbook
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import cm
from io import BytesIO
from PIL import Image as PILImage

def excel_to_pdf(excel_path, pdf_path):
    """
    指定したExcel(.xlsx)ファイル内の全シートを読み取り、
    A3・横向きのPDFにデータ(セル値)と画像を出力するサンプル関数
    """
    
    # 1) Excelファイルを読み込む (data_only=True : 数式セルを値として取得)
    workbook = load_workbook(excel_path, data_only=True)
    
    # 2) PDFドキュメントをReportLabで作成する準備 (A3を横向きで指定)
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A3)
    )
    
    # PDF内に追加していく「ストーリー」を格納するリスト
    story = []
    
    # 標準的なスタイルセットを取得 (タイトルや本文などの基本的なスタイル)
    styles = getSampleStyleSheet()
    
    # 3) ブック内のすべてのシートを順次処理
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # --- (A) シート名を見出しとして配置 ---
        heading = Paragraph(f"Sheet: {sheet_name}", styles["Heading2"])
        story.append(heading)
        story.append(Spacer(1, 0.5 * cm))  # 少し余白
        
        # --- (B) シートのセルをTableとして取得 ---
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        data = []
        for row in sheet.iter_rows(min_row=1, max_row=max_row,
                                   min_col=1, max_col=max_col,
                                   values_only=True):
            # rowはタプルで返るのでlistに変換
            data.append(list(row))
        
        if data:
            # Tableオブジェクトに変換
            table = Table(data)
            
            # テーブルのスタイルを指定(罫線、文字サイズなど)
            table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTNAME', (0,0), (-1,-1), 'Helvetica'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 0.5 * cm))
        
        # --- (C) シート内の画像を取得して配置 ---
        # 注意: openpyxlのバージョンによって画像情報の取得方法が異なる
        # 通常は sheet._images から非公開属性として取得
        for img_obj in getattr(sheet, '_images', []):
            # 画像のバイナリデータ
            # openpyxlのバージョンにより _data() / _data プロパティのどちらかになる
            # 下記は関数形式を想定
            img_data = img_obj._data()
            
            # Pillowで読み込み (BytesIOを介す)
            pil_img = PILImage.open(BytesIO(img_data))
            
            # ReportLab向けに再度BytesIOに書き出す
            tmp_io = BytesIO()
            pil_img.save(tmp_io, format='PNG')  # PNG形式などに変換
            tmp_io.seek(0)
            
            # ReportLabのImageオブジェクトとして配置
            rl_img = Image(tmp_io)
            
            # 必要に応じてサイズを制限 (例: 幅を20cmにし、縦横比は維持)
            desired_width = 20 * cm
            ratio = desired_width / pil_img.width
            rl_img._restrictSize(desired_width, pil_img.height * ratio)
            
            story.append(rl_img)
            story.append(Spacer(1, 0.5 * cm))
        
        # シートごとの区切りとして余白を挿入
        story.append(Spacer(1, 1 * cm))
    
    # 4) すべての要素をまとめてPDF出力
    doc.build(story)
    print(f"PDF 生成完了: {pdf_path}")


if __name__ == "__main__":
    excel_file = "sample_procedure.xlsx"  # 入力Excelファイルパス
    pdf_file = "output_A3.pdf"            # 出力PDFファイルパス
    
    excel_to_pdf(excel_file, pdf_file)

