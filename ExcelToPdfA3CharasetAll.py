# -*- coding: utf-8 -*-
import sys
import os
import glob
from io import BytesIO

# WindowsコンソールでUTF-8出力したい場合（Shift_JIS文字化け対策）
# （必要ならコメントアウトを外す）
# sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.utils import coordinate_from_string, column_index_from_string

from PIL import Image as PILImage

from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle,
    KeepInFrame
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import A3, landscape
from reportlab.lib.units import cm

# ▼ 日本語フォントを使う場合（Windowsの例: MSゴシック）
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
pdfmetrics.registerFont(TTFont('MS-Gothic', 'msgothic.ttc'))


def excel_to_pdf_rowwise(excel_path, pdf_path):
    """
    指定Excelファイル(全シート)を読み取り、
    A3横向きでPDF出力する。
      - 各行を1つのテーブルとして追加（「画像と分離しない」工夫）
      - 行ごとに画像があれば直後に配置
      - テーブル(行)や画像が大きい場合は KeepInFrame(mode='shrink') で縮小
      - Shift_JIS(Windows)問題は不要だが、コンソール出力で文字化けする場合はUTF-8化を推奨
    """
    
    # 1) Excelファイル読み込み (data_only=True で数式は値として取得)
    wb = load_workbook(excel_path, data_only=True)
    
    # 2) PDFドキュメント (A3横向き)
    doc = SimpleDocTemplate(
        pdf_path,
        pagesize=landscape(A3)
    )
    story = []
    
    # 3) スタイルシートを取得、日本語フォント設定
    styles = getSampleStyleSheet()
    styles["Normal"].fontName = 'MS-Gothic'
    styles["Heading2"].fontName = 'MS-Gothic'
    
    # 4) シートごとの処理
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        
        # --- シート見出し ---
        p = Paragraph(f"シート名: {sheet_name}", styles["Heading2"])
        story.append(p)
        story.append(Spacer(1, 0.5*cm))
        
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        # --- 画像アンカーを取得 (画像を「どの行」に置くか) ---
        images_in_row = {}
        for img_obj in getattr(sheet, "_images", []):
            anchor = img_obj.anchor  # 例: "C5" など
            if isinstance(anchor, str):
                # "C5" 形式の場合
                coord = coordinate_from_string(anchor)  # ("C", "5")
                rowx = int(coord[1])  # 5
            else:
                # anchorがOneCellAnchorなどのオブジェクトの場合
                rowx = anchor._from.row_idx + 1
            
            images_in_row.setdefault(rowx, []).append(img_obj)
        
        # --- 各行を順番に処理し、行データ + 画像を配置 ---
        for r in range(1, max_row + 1):
            # (A) 行データ取得
            row_values = []
            for c in range(1, max_col + 1):
                val = sheet.cell(row=r, column=c).value
                row_values.append("" if val is None else str(val))
            
            # (B) 1行のテーブル作成
            row_table = Table([row_values])
            row_table.setStyle(TableStyle([
                ('GRID', (0,0), (-1,-1), 1, colors.black),
                ('FONTNAME', (0,0), (-1,-1), 'MS-Gothic'),
                ('FONTSIZE', (0,0), (-1,-1), 8),
            ]))
            
            # KeepInFrameで囲み、フレーム内に収まらない場合は縮小
            row_table_kif = KeepInFrame(
                maxWidth=27*cm,  # ページ余白を考慮して少し狭めに
                maxHeight=15*cm, # 適当（行が長大になりすぎないように）
                content=[row_table],
                mode='shrink'
            )
            story.append(row_table_kif)
            
            # (C) 画像があれば、行の直後に配置
            if r in images_in_row:
                for img_obj in images_in_row[r]:
                    img_data = img_obj._data()
                    pil_img = PILImage.open(BytesIO(img_data))
                    
                    tmp_io = BytesIO()
                    pil_img.save(tmp_io, format='PNG')
                    tmp_io.seek(0)
                    
                    rl_img = Image(tmp_io)
                    # 幅15cmに制限
                    desired_width = 15 * cm
                    ratio = desired_width / pil_img.width
                    rl_img._restrictSize(desired_width, pil_img.height * ratio)
                    
                    # 画像自体をKeepInFrameでさらに縮小保証
                    img_kif = KeepInFrame(
                        maxWidth=27*cm,
                        maxHeight=15*cm,
                        content=[rl_img],
                        mode='shrink'
                    )
                    story.append(img_kif)
            
            # 少しだけ行間を空ける
            story.append(Spacer(1, 0.3*cm))
        
        # シート間の区切り
        story.append(Spacer(1, 1*cm))
    
    # 5) PDF出力
    doc.build(story)
    print(f"[完了] {excel_path} -> {pdf_path}")


def convert_excels_in_folder(folder_path):
    """
    指定フォルダ内にある .xlsx をすべて走査し、
    同名のPDFに変換する (A3横向き, 画像とデータを行単位で出力)
    """
    folder_path = os.path.abspath(folder_path)
    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    
    if not excel_files:
        print("指定フォルダに .xlsx ファイルが見つかりません。")
        return
    
    print(f"対象ファイル数: {len(excel_files)}")
    
    for excel_file in excel_files:
        # 同じファイル名 + .pdf
        base_name = os.path.splitext(os.path.basename(excel_file))[0]
        pdf_file = base_name + ".pdf"
        pdf_path = os.path.join(folder_path, pdf_file)
        
        excel_to_pdf_rowwise(excel_file, pdf_path)


if __name__ == "__main__":
    # 例: "C:/Users/YourName/Documents/excels"
    target_folder = "C:/path/to/excels"
    convert_excels_in_folder(target_folder)


